"""
Packing list generator.
Reads order data + packing spec rules and generates Excel packing lists.
"""

import io
import math
import re
import logging
from collections import defaultdict
from functools import reduce
from typing import Optional
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

HKM_BOX_BASE = '59*39'
HKM_DEFAULT_BOX_HEIGHT = 30
HKM_VOLUME_MAP = {10: 0.02301, 15: 0.034515, 20: 0.04602, 25: 0.057525, 30: 0.06903}
HKM_DEFAULT_VOLUME = 0.06903

THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin'),
)
HEADER_FILL = PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid')
HEADER_FONT = Font(bold=True, size=10)
NORMAL_FONT = Font(size=10)
CENTER_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)

BRA_SIZE_PATTERN = re.compile(r'^([A-H])(\d{2,3})$')
CLOTHING_SIZES = ['3XL', '2XL', 'XXL', 'XL', 'XXS', 'XS', 'S', 'M', 'L']
CUP_ORDER = ['C', 'D', 'E', 'F', 'G', 'H']
CLOTHING_SIZE_ORDER = ['XXS', 'XS', 'S', 'M', 'L', 'XL', 'XXL', '2XL', '3XL']

COL_COUNT = 24


# ──────── General Helpers ────────

def _gcd(a: int, b: int) -> int:
    while b:
        a, b = b, a % b
    return a


def _gcd_list(numbers: list[int]) -> int:
    if not numbers:
        return 1
    return reduce(_gcd, numbers)


def _is_bra_size(size: str) -> bool:
    return bool(BRA_SIZE_PATTERN.match(str(size).strip())) if size else False


def _parse_bra_size(size: str) -> tuple[str, int]:
    m = BRA_SIZE_PATTERN.match(str(size).strip())
    if m:
        return m.group(1), int(m.group(2))
    return '', 0


def _clothing_sort_key(size: str) -> int:
    s = str(size).strip().upper()
    if s in CLOTHING_SIZE_ORDER:
        return CLOTHING_SIZE_ORDER.index(s)
    return 100


def _style_cell(ws, row, col, value=None, bold=False, fill=None):
    cell = ws.cell(row=row, column=col, value=value)
    cell.border = THIN_BORDER
    cell.alignment = CENTER_ALIGN
    cell.font = HEADER_FONT if bold else NORMAL_FONT
    if fill:
        cell.fill = fill
    return cell


def _is_prepack_line(line: dict) -> bool:
    """Detect prepack lines: no size AND no colour, Description starts with a number."""
    size = (line.get('size') or '').strip()
    colour = (line.get('colour') or '').strip()
    if size or colour:
        return False
    desc = (line.get('description') or '').strip()
    return bool(re.match(r'^\d+\s+', desc))


def _extract_pack_count(description: str) -> int:
    """Extract per-pack count from prepack Description like '6 Ocean Bloom bs PP'."""
    m = re.match(r'^(\d+)\s+', (description or '').strip())
    return int(m.group(1)) if m else 1


def _build_dc_display(lines: list[dict]) -> str:
    """Build DC display: DC/Warehouse/Destination/Flow."""
    for ln in lines:
        parts = []
        for field in ('dc', 'warehouse', 'destination', 'flow'):
            val = (ln.get(field) or '').strip()
            if val:
                parts.append(val)
        if parts:
            return '/'.join(parts)
    return ''


def _get_height_specs(customer_specs: list[dict], article_desc: str) -> list[tuple[int, int, str, float]]:
    """Get (height, max_per_box, box_carton, volume) list from customer specs.

    Tries to match spec article_no keywords against article_desc.
    Returns list sorted by height ascending.
    """
    if not customer_specs:
        return []

    desc_upper = (article_desc or '').upper()
    matched = []
    for spec in customer_specs:
        article_no = (spec.get('article_no') or '').strip().upper()
        if not article_no:
            continue
        if article_no in desc_upper:
            h = spec.get('box_height') or 30
            m = spec.get('box_max') or 50
            c = spec.get('box_carton') or f'{HKM_BOX_BASE}*{h}'
            v = spec.get('box_volume') or HKM_VOLUME_MAP.get(h, HKM_DEFAULT_VOLUME)
            matched.append((h, m, c, v))

    if not matched:
        for spec in customer_specs:
            article_no = (spec.get('article_no') or '').strip()
            if not article_no:
                h = spec.get('box_height') or 30
                m = spec.get('box_max') or 50
                c = spec.get('box_carton') or f'{HKM_BOX_BASE}*{h}'
                v = spec.get('box_volume') or HKM_VOLUME_MAP.get(h, HKM_DEFAULT_VOLUME)
                matched.append((h, m, c, v))

    return sorted(matched, key=lambda x: x[0])


def _find_full_box_spec(height_specs: list, is_pa: bool) -> tuple[int, int, str, float]:
    """Get the box spec for full boxes (largest height within allowed range)."""
    min_h = 20 if is_pa else 10
    valid = [(h, m, c, v) for h, m, c, v in height_specs if h >= min_h]
    if not valid:
        valid = height_specs
    if not valid:
        h = HKM_DEFAULT_BOX_HEIGHT
        return h, 50, f'{HKM_BOX_BASE}*{h}', HKM_VOLUME_MAP.get(h, HKM_DEFAULT_VOLUME)
    return valid[-1]


def _find_remainder_box_spec(qty: int, height_specs: list, is_pa: bool) -> tuple[int, int, str, float]:
    """Find the smallest box height that fits the remaining qty."""
    min_h = 20 if is_pa else 10
    valid = [(h, m, c, v) for h, m, c, v in height_specs if h >= min_h]
    if not valid:
        valid = height_specs
    if not valid:
        h = HKM_DEFAULT_BOX_HEIGHT
        return h, 50, f'{HKM_BOX_BASE}*{h}', HKM_VOLUME_MAP.get(h, HKM_DEFAULT_VOLUME)
    for h, m, c, v in valid:
        if m >= qty:
            return h, m, c, v
    return valid[-1]


# ──────── Table Headers ────────

def _write_prepack_header_rows(ws, row_num: int):
    """Write the two header rows for prepack (配比) section."""
    header_row1 = {
        1: '箱号', 4: '箱数', 5: '款号', 6: '尺码',
        7: '每   箱   尺   码   搭   配',
        13: '每包数量', 14: '每箱包数',
        15: '合计', 16: '每箱净重', 17: '每箱毛重',
        19: '纸箱尺寸', 20: '体积', 21: '胶袋尺寸',
        22: '数量', 23: '备注',
    }
    for col_idx in range(1, COL_COUNT + 1):
        val = header_row1.get(col_idx)
        _style_cell(ws, row_num, col_idx, val, bold=True, fill=HEADER_FILL)
    ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=3)
    ws.merge_cells(start_row=row_num, start_column=7, end_row=row_num, end_column=12)

    header_row2 = {
        1: '开始箱号', 3: '结束箱号',
        7: 'C', 8: 'D', 9: 'E', 10: 'F', 11: 'G', 12: 'H',
        16: 'N.W.(kg)', 17: 'G.W.(kg)',
    }
    for col_idx in range(1, COL_COUNT + 1):
        val = header_row2.get(col_idx)
        _style_cell(ws, row_num + 1, col_idx, val, bold=True, fill=HEADER_FILL)

    return row_num + 2


def _write_detail_header_rows(ws, row_num: int):
    """Write the two header rows for detail (明细) section."""
    header_row1 = {
        1: '箱号', 4: '箱数', 5: '颜色', 6: '尺码',
        7: '每箱数量',
        15: '合计', 16: '每箱净重', 17: '每箱毛重',
        19: '纸箱尺寸', 20: '体积', 21: '胶袋尺寸',
        22: '数量', 23: '备注',
    }
    for col_idx in range(1, COL_COUNT + 1):
        val = header_row1.get(col_idx)
        _style_cell(ws, row_num, col_idx, val, bold=True, fill=HEADER_FILL)
    ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=3)

    header_row2 = {1: '开始箱号', 3: '结束箱号', 16: 'N.W.(kg)', 17: 'G.W.(kg)'}
    for col_idx in range(1, COL_COUNT + 1):
        val = header_row2.get(col_idx)
        _style_cell(ws, row_num + 1, col_idx, val, bold=True, fill=HEADER_FILL)

    return row_num + 2


def _write_group_info(ws, row_num: int, style_name: str, dc_display: str, barcode: str) -> int:
    """Write the 款号/DC/条形码 group header rows."""
    ws.cell(row=row_num, column=1, value='款号：').font = Font(bold=True, size=10)
    ws.cell(row=row_num, column=2, value=style_name)
    row_num += 1

    ws.cell(row=row_num, column=1, value='DC：').font = Font(bold=True, size=10)
    ws.cell(row=row_num, column=2, value=dc_display)
    row_num += 1

    ws.cell(row=row_num, column=1, value='条形码：').font = Font(bold=True, size=10)
    ws.cell(row=row_num, column=2, value=barcode)
    row_num += 1

    return row_num + 1


def _set_column_widths(ws):
    """Apply column widths to worksheet."""
    col_widths = {
        1: 10, 2: 4, 3: 10, 4: 8, 5: 12, 6: 8,
        7: 10, 8: 5, 9: 5, 10: 5, 11: 5, 12: 5,
        13: 10, 14: 10, 15: 10, 16: 10, 17: 10, 18: 4,
        19: 12, 20: 10, 21: 10, 22: 8, 23: 10, 24: 4,
    }
    for col_idx, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width


# ──────── HKM Combined Generator ────────

def generate_hkm_packing_list_combined(
    orders: list[dict],
    lines_per_order: list[list[dict]],
    customer_specs: Optional[list[dict]] = None,
) -> bytes:
    """Generate HKM packing list Excel.

    If prepack lines exist: write prepack section first, then detail section.
    If no prepack: write detail section only.
    Box numbering is continuous across both sections.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = '合并表'

    all_lines = []
    for order, lines in zip(orders, lines_per_order):
        order_country = (order.get('country') or '').strip()
        for ln in lines:
            ln['_country'] = order_country
        all_lines.extend(lines)

    prepack_lines = [ln for ln in all_lines if _is_prepack_line(ln)]
    detail_lines = [ln for ln in all_lines if not _is_prepack_line(ln)]

    row_num = 1
    box_start = 1

    # ── Prepack Section (配比数据) ──
    if prepack_lines:
        dest_groups = defaultdict(list)
        for ln in prepack_lines:
            dest = (ln.get('destination') or '').strip() or 'DEFAULT'
            dest_groups[dest].append(ln)

        for dest_key, group_lines in dest_groups.items():
            descriptions = []
            barcodes = []
            for ln in group_lines:
                desc = (ln.get('description') or '').strip()
                if desc and desc not in descriptions:
                    descriptions.append(desc)
                ean = (ln.get('ean') or '').strip()
                if ean and ean not in barcodes:
                    barcodes.append(ean)

            dc_display = _build_dc_display(group_lines)

            row_num = _write_group_info(
                ws, row_num,
                style_name=' / '.join(descriptions),
                dc_display=dc_display,
                barcode=' / '.join(barcodes),
            )

            row_num = _write_prepack_header_rows(ws, row_num)

            for ln in group_lines:
                qty = ln.get('quantity') or ln.get('tot_pieces') or 0
                if not qty or qty <= 0:
                    continue

                pack_count = _extract_pack_count(ln.get('description') or '')
                box_end = box_start + qty - 1
                total_pieces = pack_count * qty
                box_carton = f'{HKM_BOX_BASE}*{HKM_DEFAULT_BOX_HEIGHT}'
                volume = HKM_VOLUME_MAP.get(HKM_DEFAULT_BOX_HEIGHT, HKM_DEFAULT_VOLUME)

                row_data = [None] * COL_COUNT
                row_data[0] = box_start
                row_data[2] = box_end
                row_data[3] = qty
                row_data[12] = pack_count
                row_data[14] = total_pieces
                row_data[18] = box_carton
                row_data[19] = round(volume * qty, 5)

                for c in range(COL_COUNT):
                    _style_cell(ws, row_num, c + 1, row_data[c])
                row_num += 1
                box_start = box_end + 1

            row_num += 2

    # ── Detail Section (装箱明细) ──
    if detail_lines:
        dest_groups = defaultdict(list)
        for ln in detail_lines:
            dest = (ln.get('destination') or '').strip() or 'DEFAULT'
            dest_groups[dest].append(ln)

        for dest_key, group_lines in dest_groups.items():
            is_pa = dest_key.strip().upper() == 'PA'

            generic_no = ''
            for ln in group_lines:
                gn = (ln.get('generic_article_no') or '').strip()
                if gn:
                    generic_no = gn
                    break

            dc_display = _build_dc_display(group_lines)

            row_num = _write_group_info(
                ws, row_num,
                style_name=generic_no,
                dc_display=dc_display,
                barcode='',
            )

            row_num = _write_detail_header_rows(ws, row_num)

            remainders = []

            for ln in group_lines:
                qty = ln.get('quantity') or ln.get('tot_pieces') or 0
                if not qty or qty <= 0:
                    continue

                colour = ln.get('colour') or ''
                size = ln.get('size') or ''
                art_desc = (ln.get('article_description') or '').upper()

                height_specs = _get_height_specs(customer_specs or [], art_desc)

                if height_specs:
                    _, box_max, box_carton, box_vol = _find_full_box_spec(height_specs, is_pa)
                else:
                    box_max = 50
                    bh = HKM_DEFAULT_BOX_HEIGHT
                    box_carton = f'{HKM_BOX_BASE}*{bh}'
                    box_vol = HKM_VOLUME_MAP.get(bh, HKM_DEFAULT_VOLUME)

                num_full = qty // box_max
                rem = qty % box_max

                if num_full > 0:
                    box_end = box_start + num_full - 1
                    total = num_full * box_max

                    row_data = [None] * COL_COUNT
                    row_data[0] = box_start
                    row_data[2] = box_end
                    row_data[3] = num_full
                    row_data[4] = colour
                    row_data[5] = size
                    row_data[6] = box_max
                    row_data[14] = total
                    row_data[18] = box_carton
                    row_data[19] = box_vol

                    for c in range(COL_COUNT):
                        _style_cell(ws, row_num, c + 1, row_data[c])
                    row_num += 1
                    box_start = box_end + 1

                if rem > 0:
                    if is_pa:
                        rem_h, _, rem_carton, rem_vol = _find_remainder_box_spec(
                            rem, height_specs, is_pa) if height_specs else (
                            HKM_DEFAULT_BOX_HEIGHT, 50,
                            f'{HKM_BOX_BASE}*{HKM_DEFAULT_BOX_HEIGHT}', HKM_DEFAULT_VOLUME)

                        row_data = [None] * COL_COUNT
                        row_data[0] = box_start
                        row_data[2] = box_start
                        row_data[3] = 1
                        row_data[4] = colour
                        row_data[5] = size
                        row_data[6] = rem
                        row_data[14] = rem
                        row_data[18] = rem_carton
                        row_data[19] = rem_vol

                        for c in range(COL_COUNT):
                            _style_cell(ws, row_num, c + 1, row_data[c])
                        row_num += 1
                        box_start += 1
                    else:
                        remainders.append({
                            'colour': colour,
                            'size': size,
                            'qty': rem,
                            'height_specs': height_specs,
                        })

            # Mixed packing for non-PA remainders
            if remainders and not is_pa:
                all_max = []
                for r in remainders:
                    if r['height_specs']:
                        _, bm, _, _ = _find_full_box_spec(r['height_specs'], False)
                        all_max.append(bm)
                mixed_max = min(all_max) if all_max else 50

                current_box_items: list[dict] = []
                current_qty = 0
                mixed_boxes: list[list[dict]] = []

                for item in remainders:
                    if current_qty + item['qty'] <= mixed_max:
                        current_box_items.append(item)
                        current_qty += item['qty']
                    else:
                        if current_box_items:
                            mixed_boxes.append(current_box_items)
                        current_box_items = [item]
                        current_qty = item['qty']
                if current_box_items:
                    mixed_boxes.append(current_box_items)

                for box_items in mixed_boxes:
                    total_qty = sum(i['qty'] for i in box_items)

                    specs_to_check = box_items[0]['height_specs'] if box_items[0]['height_specs'] else []
                    if specs_to_check:
                        _, _, mix_carton, mix_vol = _find_remainder_box_spec(
                            total_qty, specs_to_check, False)
                    else:
                        bh = HKM_DEFAULT_BOX_HEIGHT
                        mix_carton = f'{HKM_BOX_BASE}*{bh}'
                        mix_vol = HKM_VOLUME_MAP.get(bh, HKM_DEFAULT_VOLUME)

                    first_item = True
                    for item in box_items:
                        row_data = [None] * COL_COUNT
                        if first_item:
                            row_data[0] = box_start
                            row_data[2] = box_start
                            row_data[3] = 1
                            row_data[18] = mix_carton
                            row_data[19] = mix_vol
                            first_item = False
                        row_data[4] = item['colour']
                        row_data[5] = item['size']
                        row_data[6] = item['qty']
                        row_data[14] = item['qty']

                        for c in range(COL_COUNT):
                            _style_cell(ws, row_num, c + 1, row_data[c])
                        row_num += 1
                    box_start += 1

            row_num += 2

    _set_column_widths(ws)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def generate_hkm_packing_list(
    order: dict,
    lines: list[dict],
    customer_specs: Optional[list[dict]] = None,
) -> bytes:
    """Generate HKM packing list for a single order (backward compatible)."""
    return generate_hkm_packing_list_combined([order], [lines], customer_specs)


# ──────── Generic Packing List Generator ────────

def generate_generic_packing_list(
    order: dict,
    lines: list[dict],
    customer_specs: Optional[list[dict]] = None,
) -> bytes:
    """Generate a generic (non-HKM) packing list Excel.

    Uses PackingSpec rules, box height auto-matching, and mixed packing.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = '装箱单'

    po_number = order.get('po', '')
    generic_article_no = order.get('generic_article_no', '')
    country = (order.get('country') or '').upper()
    article_description = (order.get('article_description') or '').upper()

    ws.append(['订单号', po_number])
    ws.append(['客款号', generic_article_no])
    ws.append(['国家', country])
    ws.append([])

    headers = [
        '箱号(开始)', '箱号(结束)', '箱数', '款式图', '厂款号',
        '颜色', '尺码', '每箱数量', '合计',
        '每箱净重', '每箱毛重', '纸箱尺寸', '体积', '备注',
    ]

    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=5, column=i, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = CENTER_ALIGN

    row_num = 6
    box_start = 1

    default_max = 50
    default_box = '60*40*25'
    default_volume = 0.06

    height_specs = _get_height_specs(customer_specs or [], article_description)

    if height_specs:
        _, default_max, default_box, default_volume = height_specs[-1]
    elif customer_specs:
        spec = customer_specs[0]
        default_max = spec.get('box_max') or 50
        default_box = spec.get('box_carton') or default_box
        default_volume = spec.get('box_volume') or default_volume

    dest_groups = defaultdict(list)
    for ln in lines:
        dest = (ln.get('destination') or '').strip() or 'DEFAULT'
        dest_groups[dest].append(ln)

    for dest_key, group_lines in dest_groups.items():
        is_pa = dest_key.strip().upper() == 'PA'
        remainders = []

        for line in group_lines:
            qty = line.get('quantity') or line.get('tot_pieces') or 0
            if not qty or qty <= 0:
                continue

            line_desc = (line.get('article_description') or '').upper()
            line_height_specs = _get_height_specs(customer_specs or [], line_desc) or height_specs

            if line_height_specs:
                _, box_max, box_carton, box_vol = _find_full_box_spec(line_height_specs, is_pa)
            else:
                box_max = default_max
                box_carton = default_box
                box_vol = default_volume

            num_full = qty // box_max
            rem = qty % box_max

            if num_full > 0:
                box_end = box_start + num_full - 1
                total = num_full * box_max
                _style_cell(ws, row_num, 1, box_start)
                _style_cell(ws, row_num, 2, box_end)
                _style_cell(ws, row_num, 3, num_full)
                _style_cell(ws, row_num, 5, generic_article_no)
                _style_cell(ws, row_num, 6, line.get('colour', ''))
                _style_cell(ws, row_num, 7, line.get('size', ''))
                _style_cell(ws, row_num, 8, box_max)
                _style_cell(ws, row_num, 9, total)
                _style_cell(ws, row_num, 12, box_carton)
                _style_cell(ws, row_num, 13, box_vol)
                row_num += 1
                box_start = box_end + 1

            if rem > 0:
                if is_pa:
                    if line_height_specs:
                        _, _, rem_carton, rem_vol = _find_remainder_box_spec(
                            rem, line_height_specs, True)
                    else:
                        rem_carton, rem_vol = box_carton, box_vol

                    _style_cell(ws, row_num, 1, box_start)
                    _style_cell(ws, row_num, 2, box_start)
                    _style_cell(ws, row_num, 3, 1)
                    _style_cell(ws, row_num, 5, generic_article_no)
                    _style_cell(ws, row_num, 6, line.get('colour', ''))
                    _style_cell(ws, row_num, 7, line.get('size', ''))
                    _style_cell(ws, row_num, 8, rem)
                    _style_cell(ws, row_num, 9, rem)
                    _style_cell(ws, row_num, 12, rem_carton)
                    _style_cell(ws, row_num, 13, rem_vol)
                    row_num += 1
                    box_start += 1
                else:
                    remainders.append({
                        'colour': line.get('colour', ''),
                        'size': line.get('size', ''),
                        'qty': rem,
                        'height_specs': line_height_specs,
                    })

        # Mixed packing for non-PA remainders
        if remainders and not is_pa:
            all_max = []
            for r in remainders:
                if r['height_specs']:
                    _, bm, _, _ = _find_full_box_spec(r['height_specs'], False)
                    all_max.append(bm)
            mixed_max = min(all_max) if all_max else default_max

            current_box_items: list[dict] = []
            current_qty = 0
            mixed_boxes: list[list[dict]] = []

            for item in remainders:
                if current_qty + item['qty'] <= mixed_max:
                    current_box_items.append(item)
                    current_qty += item['qty']
                else:
                    if current_box_items:
                        mixed_boxes.append(current_box_items)
                    current_box_items = [item]
                    current_qty = item['qty']
            if current_box_items:
                mixed_boxes.append(current_box_items)

            for box_items in mixed_boxes:
                total_qty = sum(i['qty'] for i in box_items)
                specs_to_check = box_items[0]['height_specs']
                if specs_to_check:
                    _, _, mix_carton, mix_vol = _find_remainder_box_spec(
                        total_qty, specs_to_check, False)
                else:
                    mix_carton, mix_vol = default_box, default_volume

                for bi, item in enumerate(box_items):
                    _style_cell(ws, row_num, 1, box_start if bi == 0 else None)
                    _style_cell(ws, row_num, 2, box_start if bi == 0 else None)
                    _style_cell(ws, row_num, 3, 1 if bi == 0 else None)
                    _style_cell(ws, row_num, 5, generic_article_no)
                    _style_cell(ws, row_num, 6, item['colour'])
                    _style_cell(ws, row_num, 7, item['size'])
                    _style_cell(ws, row_num, 8, item['qty'])
                    _style_cell(ws, row_num, 9, item['qty'])
                    _style_cell(ws, row_num, 12, mix_carton if bi == 0 else None)
                    _style_cell(ws, row_num, 13, mix_vol if bi == 0 else None)
                    _style_cell(ws, row_num, 14, '混装' if bi == 0 else None)
                    row_num += 1
                box_start += 1

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
