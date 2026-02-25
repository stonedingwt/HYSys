import csv
import io
from typing import Optional

from fastapi import APIRouter, Query, UploadFile, File
from fastapi.responses import StreamingResponse

from mep.common.schemas.api import resp_200, resp_500
from mep.database.models.data_dict import DataDictDao, DictCategory

router = APIRouter(prefix='/data-dict', tags=['data_dict'])


# ─── Category endpoints ───

@router.get('/category/tree')
async def category_tree():
    try:
        tree = await DataDictDao.get_category_tree()
        return resp_200(tree)
    except Exception as e:
        return resp_500(str(e))


@router.get('/category/list')
async def category_list():
    """Flat list of all categories (for dropdowns)."""
    try:
        cats = await DataDictDao.get_all_categories()
        return resp_200([c.dict() for c in cats])
    except Exception as e:
        return resp_500(str(e))


@router.post('/category/create')
async def category_create(data: dict):
    try:
        data.pop('id', None)
        data.pop('create_time', None)
        data.pop('update_time', None)
        data.pop('children', None)
        item = await DataDictDao.create_category(data)
        return resp_200(item.dict() if item else None)
    except Exception as e:
        return resp_500(str(e))


@router.put('/category/update')
async def category_update(data: dict):
    cat_id = data.pop('id', None)
    if not cat_id:
        return resp_500('Missing id')
    try:
        data.pop('create_time', None)
        data.pop('update_time', None)
        data.pop('children', None)
        item = await DataDictDao.update_category(cat_id, data)
        return resp_200(item.dict() if item else None)
    except Exception as e:
        return resp_500(str(e))


@router.delete('/category/delete')
async def category_delete(id: int = Query(...)):
    try:
        ok = await DataDictDao.delete_category(id)
        return resp_200({'success': ok})
    except Exception as e:
        return resp_500(str(e))


# ─── Item endpoints ───

@router.get('/item/list')
async def item_list(
    category_id: int = Query(0),
    keyword: str = Query(''),
    page_num: int = Query(1, ge=1),
    page_size: int = Query(15, ge=1, le=100),
    sort_by: str = Query(''),
    sort_order: str = Query('asc'),
):
    try:
        items, total = await DataDictDao.list_items(
            category_id, keyword, page_num, page_size, sort_by, sort_order)
        return resp_200({
            'data': [i.dict() for i in items],
            'total': total,
            'page_num': page_num,
            'page_size': page_size,
        })
    except Exception as e:
        return resp_500(str(e))


@router.get('/item/tree')
async def item_tree(category_id: int = Query(...)):
    try:
        tree = await DataDictDao.get_item_tree(category_id)
        return resp_200(tree)
    except Exception as e:
        return resp_500(str(e))


@router.post('/item/create')
async def item_create(data: dict):
    try:
        data.pop('id', None)
        data.pop('create_time', None)
        data.pop('update_time', None)
        item = await DataDictDao.create_item(data)
        return resp_200(item.dict() if item else None)
    except Exception as e:
        return resp_500(str(e))


@router.put('/item/update')
async def item_update(data: dict):
    item_id = data.pop('id', None)
    if not item_id:
        return resp_500('Missing id')
    try:
        data.pop('create_time', None)
        data.pop('update_time', None)
        item = await DataDictDao.update_item(item_id, data)
        return resp_200(item.dict() if item else None)
    except Exception as e:
        return resp_500(str(e))


@router.delete('/item/delete')
async def item_delete(id: int = Query(...)):
    try:
        ok = await DataDictDao.delete_item(id)
        return resp_200({'success': ok})
    except Exception as e:
        return resp_500(str(e))


# ─── Import ───

def _parse_csv(content: bytes) -> list:
    text = content.decode('utf-8-sig')
    reader = csv.DictReader(io.StringIO(text))
    return [dict(row) for row in reader]


def _parse_excel(content: bytes) -> list:
    from openpyxl import load_workbook
    wb = load_workbook(filename=io.BytesIO(content), read_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    header = [str(h or '').strip().lower() for h in next(rows_iter)]
    result = []
    for row in rows_iter:
        d = {}
        for i, val in enumerate(row):
            if i < len(header):
                d[header[i]] = str(val).strip() if val is not None else ''
        result.append(d)
    wb.close()
    return result


@router.post('/import')
async def import_data(file: UploadFile = File(...)):
    try:
        content = await file.read()
        filename = (file.filename or '').lower()

        if filename.endswith('.xlsx') or filename.endswith('.xls'):
            rows = _parse_excel(content)
        elif filename.endswith('.csv'):
            rows = _parse_csv(content)
        else:
            return resp_500('Unsupported file type. Please use .csv or .xlsx')

        if not rows:
            return resp_500('File is empty or has no data rows')

        result = await DataDictDao.batch_import(rows)
        return resp_200(result)
    except Exception as e:
        return resp_500(str(e))


# ─── Export ───

@router.get('/export')
async def export_data(
    category_id: int = Query(0),
    keyword: str = Query(''),
):
    try:
        items = await DataDictDao.list_all_items(category_id, keyword)
        cats = await DataDictDao.get_all_categories()
        cat_map = {c.id: c for c in cats}

        output = io.StringIO()
        output.write('\ufeff')
        writer = csv.writer(output)
        writer.writerow(['cat_code', 'cat_name', 'item_label', 'item_value', 'sort_order', 'status', 'remark'])

        for item in items:
            cat = cat_map.get(item.category_id)
            writer.writerow([
                cat.cat_code if cat else '',
                cat.cat_name if cat else '',
                item.item_label,
                item.item_value,
                item.sort_order,
                item.status,
                item.remark or '',
            ])

        output.seek(0)
        return StreamingResponse(
            iter([output.getvalue()]),
            media_type='text/csv; charset=utf-8',
            headers={'Content-Disposition': 'attachment; filename=data_dict.csv'}
        )
    except Exception as e:
        return resp_500(str(e))
